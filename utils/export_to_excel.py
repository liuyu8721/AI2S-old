# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
import json

import openpyxl
import re
import os
import time
from LAC import LAC


def original_data(file_path):
    """

    :return: 返回的是经过对不规则符号进行处理之后的数据
    """
    s = []
    s1 = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    for i in range(2, sheet.max_row + 1):
        s.append(sheet['C' + str(i)].value)
    return s


def data_handle(filepath):
    if not os.path.exists('./海南疫情.xlsx'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '数据'
        sheet.append(
            ('原文', '病例编号', '性别', '年龄', '常住地', '关系', '日期', '句子', '交通编号', '交通方式', '出发点', '目的地'))
    else:
        wb = openpyxl.load_workbook('./海南疫情.xlsx')
        sheet = wb.worksheets[0]
    s3 = []
    s4 = []
    list1 = []
    s1 = original_data(filepath)
    data_list = []
    for i in s1:
        flag = False
        routes = []
        nums = re.search(r'第\d+', i)
        if nums:
            num = nums.group()
        else:
            num = ''
        sex = re.search(r'[男|女]', i)  # 正则得到性别
        ages = re.search(r'\d+[\u5c81]', i)  # 正则得到年龄
        citys = re.search(r'常住\w+', i)
        if citys:
            city = citys.group()
        else:
            city = ''
        relation1 = re.search(r'(\d+号\、?\w+病例朋友)', i)
        relation2 = re.search(r'(系\d[\w+\、?]+)|(与第\w+[\、?\w+]+)(亲属关系|亲友关系|同一小区|一起乘网约车)|(与确诊患者刘XX临高2号病例同乘私家车)', i)
        if relation2 and relation1:
            s = relation2.group() + ',' + relation1.group()
            flag = True
        if relation2:
            if flag:
                relation = s
            else:
                relation = relation2.group()
        else:
            relation = ''
        s3 = (i.split('。'))
        for j in s3:
            if j == '\n' or j == '':
                continue
            else:
                time = re.findall(r'\d+月\d+日', j)
                if time != []:
                    for t in time:
                        list1 = []
                        dict_y = []
                        sentence = j.split(t)[1]
                        times = re.search(r'\d+月\d+日', sentence)
                        sentences = re.findall(
                            r'(乘\w+|转运至\w+|去\w+|到\w+|到达\w+|从\w+|在\w+|进入\w+|返回\w+|抵达\w+|转至\w+|入住\w+|前往\w+|送至\w+|接往\w+|转至\w+|乘坐\w+)',
                            sentence)
                        if times and len(sentences) > 1:
                            list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence.split('，')[0]])
                            if sentence != []:
                                sheet.append(process(sentence, list1))
                        elif 1 < len(sentences):
                            for z in range(len(sentences)):
                                list1 = []
                                list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence])
                                if sentence[z] != []:
                                    rest = process(sentences[z], list1).copy()
                                    sheet.append(rest)
                        else:
                            list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence])
                            if sentence != []:
                                sheet.append(process(sentence, list1))
    wb.save('./海南疫情.xlsx')


def process(lines, list1):
    """
    Args:
        in_text (_type_): _description_
        out_text (_type_): _description_
    """
    # 装载LAC模型
    lac = LAC(mode='lac')
    s = re.search(r'机场', lines)
    s1 = re.search(r'市妇幼保健院', lines)
    s2 = re.search(r'琼海市博鳌镇幸福海小区家中', lines)
    s3 = re.search(r'市人民医院', lines)
    s4 = re.search(r'大小洞天', lines)
    s5 = re.search('r天涯区升升超市', lines)
    s6 = re.search(r'第一市场河西路口好芙利蛋糕店', lines)
    s7 = re.search(r'川味饭店', lines)
    s8 = re.search(r'海鲜广场', lines)
    s9 = re.search(r'一心堂药店', lines)
    s10 = re.search(r'群众街', lines)
    s11 = re.search(r'川味王餐厅', lines)
    s12 = re.search(r'羊栏市场', lines)
    s13 = re.search(r'县人民医院', lines)
    s14 = re.search(r'省人民医院', lines)
    s15 = re.search(r'临高县碧桂园金沙滩南享养生会所', lines)
    s16 = re.search(r'南享养生会所', lines)
    s17 = re.search(r'省定点医院', lines)
    s18 = re.search(r'临高县临城镇工会', lines)
    s19 = re.search(r'海口市美兰区三江农场派出所', lines)
    s20 = re.search(r'新媒体绿都小区', lines)
    s21 = re.search(r'美兰机场', lines)
    s22 = re.search(r'临高县半岛阳光小区', lines)
    s23 = re.search(r'三亚凤凰机场', lines)
    s24 = re.search(r'湖北襄阳机场', lines)
    s25 = re.search(r'明珠商业城百佳汇超市', lines)
    s26 = re.search(r'君澜三亚湾迎宾馆', lines)
    s27 = re.search(r'海口琼山大园', lines)
    s28 = re.search(r'文昌公园', lines)
    s29 = re.search(r'杏林小区旁瓜菜批发市场', lines)
    s30 = re.search(r'明珠商业城百佳汇超市', lines)
    s31 = re.search(r'海安港', lines)
    s32 = re.search(r'三亚市吉阳区君和君泰小区', lines)
    s33 = re.search(r'三亚市崖州区南滨佳康宾馆', lines)
    s34 = re.search(r'中西结合医院', lines)
    s35 = re.search(r'石碌镇城区', lines)
    s36 = re.search(r'昌化镇昌城村', lines)
    s37 = re.search(r'三亚定点留观点海角之旅酒店', lines)
    s38 = re.search(r'万宁市石梅湾九里三期', lines)
    s39 = re.search(r'三亚市吉阳区凤凰路南方航空城', lines)
    s40 = re.search(r'白马井镇海花岛', lines)
    lac_result = lac.run(lines)

    my_list = ['天涯海角景区', '海安', '旧港', '海安', '新港', '八爪鱼超市', '天涯区', '海口市', '板桥', '土福湾', '顺泽福湾', '亚龙湾壹号', '百花谷',
               '亚龙湾', '红塘湾建国酒店', '国光毫生度假酒店', '海棠区',
               '南田农场', '广西村', '水蛟村委会大园村', '新城路鲁能三亚湾升升超市', '金中海蓝钻小区', '解放路', '旺豪超市',
               '南海', '嘉园', '广西自治区', '黄姚古镇', '一心堂', '大广岛药品超市', '钟廖村', '山根镇', '水央屈村', '万宁市',
               '神州半岛', '君临海小区', '万城镇山泉海小区', '崖州区', '南海', '嘉园', '湛江市', '徐闻', '海安港', '南滨佳康宾馆'
        , '崖州区', '东合逸海郡', '云海台小区', '崖城怡宜佳超市', '吉阳区', '双大山湖湾小区', '星域小区', '神州租车公司', '水蛟村'
        , '武汉市', '武汉市', '汉口站', '广西', '南宁', '北海', '义龙西路汉庭酒店新温泉', '乐东', '千家镇'
        , '只文村', '临城路', '亿佳超市', '君澜三亚湾', '迎宾馆', '儋州市', '光村镇', '文昌市', '月亮城', '新海港', '琼海市', '琼海', '嘉积镇', '天来泉二区',
               '陵水县', '陵水', '英州镇', '椰田古寨', '兴隆镇', '曼特宁温泉', '神州半岛', '君临海小区', '神州半岛', '新月海岸小区'
        , '临高县', '碧桂园', '金沙滩', '浪琴湾', '儋州市', '光村镇', '解放军总医院', '海南医院', '儋州', '客来湾海鲜店', '新港', '龙华区', '东方市',
               '涛升国际小区', '涛昇国际小区'
        , '广东', '徐闻海安', '文昌', '东郊椰林', '秀英港', '美兰机场', '石碌镇', '昌化镇', '澄迈县', '福山镇', '昌江县', '昌化镇', '昌城村', '三亚市', '东方市',
               '八所镇', '海口鹏', '泰兴购物广场', '泰兴购物广场海甸三西路', '文化南路'
        , '金洲大厦', '金廉路广电家园', '美兰区三江农场派出所', '蓝天路', '金廉路广电家园', '秀英区', '锦地翰城', '定安县', '香江丽景小区', '石头公园'
        , '美兰区', '海口市琼山区', '金宇街道', '城西镇', '中沙路', '琼山', '灵山镇', '琼山区', '新大洲大道', '昌洒镇', '文城', '澄迈县', '老城镇', '万达广场', '湘忆亭',
               '骑楼小吃街',
               '龙桥镇', '玉良村', '老城开发区', '软件园商业步行街', '石山镇', '美鳌村', '那大镇', '儋州街', '三亚市崖州区', '南滨佳康宾馆', '水蛟村委会', '三亚旺豪超市',
               '福源小区',
               '新三得羊羔店', '石碌镇晨鑫饭店', '湖北', '湖北省', '武汉市', '湖南', '衡山', '新疆', '琼山区佳捷精品酒店', '滨江店', '江苏', '无锡',
               '福架村', '皇马假日游艇度假酒店', '坡博农贸市场', '旺佳旺超市南沙店', '凤凰路南方航空城', '迎宾路', '南山寺']
    # my_list表示需要进行拼接的地名
    my1_list = ['南山寺', '天涯海角景区', '百花谷', '亚龙湾', '威斯汀酒店', '云海台小区', '崖城怡宜佳超市', '万应堂大药房购药', '天涯区', '文昌市月亮城'
        , '千百汇超市', '乐东县人民医院', '三亚市人民医院', '君澜三亚湾迎宾馆', '三亚湾', '云海台', '崖城怡宜佳超市', '海安新港', '杏林小区旁瓜菜批发市场',
                '文昌公园', '杏林小区旁瓜菜批发市场', '明珠商业城百佳汇超市', '一心堂', '大广岛药品超市', '南滨农贸市场', '南海嘉园', '三亚市吉阳区', '海口',
                '新风街', '三亚湾', '凤翔路', '中山南路', '老城四季康城', '绿水湾小区', '天涯海角', '玫瑰谷', '福山骏群超市', '那大镇兰洋路', '中兴大道']
    # my1_list表示需要用分号隔开的地名
    location_list = []
    seg_list = lac_result[0]
    pos_list = lac_result[1]
    for i in range(len(seg_list)):
        if pos_list[i] == "LOC" or pos_list[i] == "ORG":
            location_list.append(seg_list[i])
    lines = lines.replace(' ', '')
    Tcode = re.search(
        r'(出租车|网约车|动车|私家车|旅游大巴|酒店专车|飞机|公交车|商务车|轮渡|旅游团车|滴滴专车|摩托车|救护车|急救车|海汽快车|高铁|电动车|航班|自驾|私家车|船|列车|骑车|驾车|火车|大巴|摩托三轮车|三轮车|小轿车|车|电火车|滴滴车|海汽专线|班车)',
        lines)
    T = re.search(
        '([京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼使领][A-HJ-NP-Z][A-HJ-NP-Z0-9]{4}[A-HJ-NP-Z0-9挂学警港澳])|(班次号\s?[0-9]+)|(双泰\d+号|黎母号|海棠湾号|南方\d号轮渡|紫荆\d+号|粤海铁3号|黎姆号|海装18号)|(\d+)班车|(?:车次|火车)\s?([GCDZTSPKXLY1-9]\d{1,4})|乘([GCDZTSPKXLY1-9]\d{1,4})次火车|(?:国航|南航|海航|航班|航班号|航班号是)\s?([A-Z\d]{2}\d{3,4})|([A-Z\d]{2}\d{3,4})(?:次航|航班)|(?:乘|由|被)\s?([\d]+)(?=路|救护车|急救车|车)',
        lines)
    start = ''
    end = ''
    if s != None:
        end = s.group()
    if s1 != None:
        end = s1.group()
    if s2 != None:
        end = s2.group()
    if s3 != None:
        end = s3.group()
    if s4 != None:
        end = s4.group()
    if s5 != None:
        end = s5.group()
    if s6 != None:
        end = s6.group()
    if s7 != None:
        end = s7.group()
    if s8 != None:
        end = s8.group()
    if s9 != None:
        end = s9.group()
    if s10 != None:
        end = s10.group()
    if s11 != None:
        end = s11.group()
    if s12 != None:
        end = s12.group()
    if s13 != None:
        end = s13.group()
    if s14 != None:
        end = s14.group()
    if s15 != None:
        end = s15.group()
    if s16 != None:
        end = s16.group()
    if s17 != None:
        end = s17.group()
    if s18 != None:
        end = s18.group()
    if s19 != None:
        end = s19.group()
    if s20 != None:
        end = s20.group()
    if s21 != None:
        end = s21.group()
    if s22 != None:
        end = s22.group()
    if s23 != None:
        end = s23.group()
    if s24 != None:
        end = s24.group()
    if s25 != None:
        end = s25.group()
    if s26 != None:
        end = s26.group()
    if s27 != None:
        end = s27.group()
    if s28 != None:
        end = s28.group()
    if s29 != None:
        end = s29.group()
    if s30 != None:
        end = s30.group()
    if s31 != None:
        end = s31.group()
    if s32 != None:
        end = s32.group()
    if s33 != None:
        end = s33.group()
    if s34 != None:
        end = s34.group()
    if s35 != None:
        end = s35.group()
    if s36 != None:
        end = s36.group()
    if s37 != None:
        end = s37.group()
    if s38 != None:
        end = s38.group()
    if s39 != None:
        end = s39.group()
    if s40 != None:
        end = s40.group()

    if len(location_list) == 2:
        if location_list[0] in my_list and location_list[1]:
            end = location_list[0] + location_list[1]
            start = ''
        else:
            start = location_list[0]
            end = location_list[1]
        if location_list[0] in my1_list and location_list[1] in my1_list:
            end = location_list[0] + ';' + location_list[1]
            start = ''

    if len(location_list) == 3:
        if location_list[0] in my_list and location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[0] + location_list[1] + location_list[2]
            start = ''
        elif location_list[0] in my1_list and location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[1] + location_list[2]
            start = location_list[0] + ';' + end
        elif location_list[0] in my_list and location_list[1] in my_list:
            start = location_list[0] + location_list[1]
            end = location_list[2]
        elif location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[1] + location_list[2]
            start = location_list[0]
        print('起点：', start)
        print('终点：', end)
    if T != None and Tcode:
        # 提取车牌号
        if T.group(1):
            t = T.group(1)

        # 提取班次号
        if T.group(2):
            t = T.group(2)

        # 提取特殊交通  (双泰\d+号|黎母号|海棠湾号|南方\d号轮渡|紫荆\d+号|粤海铁3号|黎姆号|海装18号)
        if T.group(3):
            t = T.group(3)

        # 提取班车号
        if T.group(4):
            t = T.group(4)

        # 提取火车有两种格式
        if T.group(5):
            t = T.group(5)
        elif T.group(6):
            t = T.group(6)

        # 提取航班号有两种格式
        if T.group(7):
            t = T.group(7)
        elif T.group(8):
            t = T.group(8)

        # 提取特殊交通 例如120
        if T.group(9):
            t = T.group(9)

        print('航班：', t)
        dict_y = [t, Tcode.group(), start, end]

        print(dict_y)
        list1.extend(dict_y)

    elif Tcode:
        dict_y = ['', Tcode.group(), start, end]
        list1.extend(dict_y)
        print(dict_y)
    elif T:

        dict_y = [T.group(), '', start, end]
        list1.extend(dict_y)
    else:
        dict_y = ['', '', start, end]
        list1.extend(dict_y)
    return list1


if __name__ == '__main__':
    data_handle('./hainan_Track15.xlsx')
