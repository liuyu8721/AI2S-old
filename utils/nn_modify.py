import openpyxl
from pymysql import *


# from openpyxl import Workbook

def all_in():
    # 导入excel文档
    wb = openpyxl.load_workbook('./7.xlsx')
    # 链接到航班数据库，host为本地主机，端口是3306，用户是root，密码是Scott624，连接上的数据库为flight，编码是utf8
    link = connect(
        host='localhost', port=3306,
        user='root', password='Scott624',
        database='flight', charset='utf8')

    # 获取游标对象，游标是系统为用户开设的一个数据缓冲区，存放SQL语句的执行结果，结果是零条、一条或多条数据
    data1 = link.cursor()
    data2 = link.cursor()
    data3 = link.cursor()
    data4 = link.cursor()
    write1 = link.cursor()
    write2 = link.cursor()

    # 获取excel文档的数据表
    ws = wb['数据']

    # 命名要存放数据的列
    ws['M1'] = '机场/火车站点'
    ws['N1'] = '火车站序'
    ws['O1'] = '出发时间'
    ws['P1'] = '到达时间'
    # 获取excel数据表的，第I列单元格
    col_i = ws['I']
    # 获取I列单元格的最大长度
    _max = len(col_i)
    # 因为第一行数据是不需要进行处理的，所以处理数据要从第二行开始
    x = 2

    # 从第九列的第二行到最后一行，读取数据，读取的数据是列车号或者航班号
    for j in ws.iter_rows(min_row=2, max_row=_max, max_col=9, min_col=9):
        for _cell in j:
            # 列车号或者航班号不为空
            if _cell.value is not None:
                # 并且该行的第10列的内容为航班时
                if ws.cell(x, 10).value == '航班':
                    # 获取该行第K、L列的数据。第K列的数据是出发地，第L列的数据为目的地
                    start_station_cell = 'K' + str(x)
                    arrive_station_cell = 'L' + str(x)
                    # 获得单元格内容
                    start_station = ws[start_station_cell]
                    arrive_station = ws[arrive_station_cell]

                    # 使用sql语句查询数据库
                    # execute语句可以执行存放SQL语句的字符串变量，或直接执行SQL语句字符串
                    # 需要查询的数据由.format来填充
                    # start_station.value是出发地，arrive_station.value[0:2]取目的地的前两个字为查询输入字段，_cell.value是航班号
                    data1.execute(
                        """
                        select StartDrome,ArriveDrome,StartTime,ArriveTime from airlinedate2
                        where startCity=(
                            select a.Abbreviation from domestic as a where a.Address_cn='{}'
                        )
                        and lastCity=(
                            select b.Abbreviation from domestic as b where b.Address_cn='{}'
                        ) and AirlineCode='{}';
                        """.format(start_station.value, arrive_station.value[0:2], _cell.value)
                    )

                    # fetchall使用游标从查询中检索
                    all_data = data1.fetchall()

                    # 如果All中的数据不为空
                    if all_data:
                        # 首先利用sql语句，将数据写入一个新表中。_cell.value是航班号，all_data[0][0]是出发机场，all_data[0][1]是降落机场，all_data[0][2]是起飞时间，all_data[0][3]是降落时间。同时要在起飞机场、降落机场和航班号都不同的情况下才写入到新的表中
                        write1.execute(
                            """
                            insert into airlinenum (Code,Start, Arrive, StartTime, ArriveTime)
                            select '{}', '{}', '{}', '{}', '{}' from DUAL
                            where not exists(select * from airlinenum where Start='{}' and Arrive='{}' and Code='{}');
                            """.format(
                                _cell.value, all_data[0][0], all_data[0][1], all_data[0][2], all_data[0][3], all_data[0][0], all_data[0][1],
                                _cell.value)
                        )
                        # 表示允许写入
                        link.commit()

                        # All中的数据，分别存放的是起飞机场、降落机场、起飞时间和降落时间

                        # 用于存放起降机场
                        row1 = 'M' + str(x)
                        # 用于存放起飞时间
                        row2 = 'O' + str(x)
                        # 用于存放降落时间
                        row3 = 'P' + str(x)

                        # 起飞机场和降落机场之间由'|'隔开
                        ws[row1] = all_data[0][0] + '|' + all_data[0][1]
                        # 写入起飞时间
                        ws[row2] = all_data[0][2]
                        # 写入降落时间
                        ws[row3] = all_data[0][3]

                    # 保存为8.xlsx
                    wb.save('8.xlsx')

                else:
                    # 用于存放列车经过的站点
                    save_station = ""
                    # 用于存放列车经过站点的站序
                    save_s_no = ""
                    # 用于存放列车的进站时间
                    save_a_time = ""
                    # 用于存放列车的出站时间
                    save_d_time = ""

                    if _cell.value is not None:
                        # 获取该行第K、L列的数据。第K列的数据是出发地，第L列的数据为目的地
                        train_start_cell = 'K' + str(x)
                        train_arrive_cell = 'L' + str(x)

                        train_start = ws[train_start_cell]
                        train_arrive = ws[train_arrive_cell]

                        # 利用sql语句进行查询。通过列车号、出发站和到达站查询途径站点、站序、进站时间和出站时间。_cell.value是列车号，train_start.value是出发站，train_arrive.value是到达站
                        data2.execute(
                            """
                            select Station,S_No,A_Time,D_Time from train
                            where ID = '{}' and S_No >= (
                                select S_No from train where ID = '{}' and Station like '{}%' limit 1
                            ) and S_No <= (
                                select S_No from train where ID = '{}' and Station like '{}%' limit 1
                            ) order by S_No;
                            """.format(_cell.value, _cell.value, train_start.value, _cell.value, train_arrive.value)
                        )

                        # result中的数据分别是站名、站序、到达时间和出发时间
                        result = data2.fetchall()

                        # 获取result的最后一个元素的索引下标
                        tlen = len(result) - 1

                        for j in range(len(result)):
                            # 将所经过的站点用'|'隔开，并在每一个站后加上一个'站'字
                            save_station = save_station + result[j][0] + '站' + '|'
                            # 将所经过的站点的站序用'|'隔开
                            save_s_no = save_s_no + str(result[j][1]) + '|'
                            # 将所经过的站点的进站时间用'|'隔开
                            save_a_time = save_a_time + result[j][2] + '|'
                            # 将所经过的站点的出站时间用'|'隔开
                            save_d_time = save_d_time + result[j][3] + '|'

                        # 将所有需要储存的数据的最后一个竖线去除
                        save_station = save_station[0:len(save_station) - 1]
                        save_s_no = save_s_no[0:len(save_s_no) - 1]
                        save_a_time = save_a_time[0:len(save_a_time) - 1]
                        save_d_time = save_d_time[0:len(save_d_time) - 1]

                        # 如果result中的存有数据
                        if result:
                            # 将列车号、出发站的站序和到达站的站序储存在新表中。_cell.value是列车号，result[0][1]是出发站的站序，result[tlen][1]是到达站的站序
                            write2.execute(
                                """
                                insert into trainnum (Code,startNum,arriveNum)
                                select '{}', '{}', '{}' from DUAL
                                where not exists(select * from trainnum
                                where Code='{}' and startNum='{}' and arriveNum='{}');
                                """.format(
                                    _cell.value, result[0][1], result[tlen][1], _cell.value, result[0][1],
                                    result[tlen][1]
                                )
                            )
                            link.commit()

                            # 储存站名
                            row1 = 'M' + str(x)
                            # 储存站序
                            row2 = 'N' + str(x)
                            # 储存到达时间
                            row3 = 'O' + str(x)
                            # 储存发车时间
                            row4 = 'P' + str(x)

                            ws[row1] = save_station
                            ws[row2] = save_s_no
                            ws[row3] = save_a_time
                            ws[row4] = save_d_time

                        wb.save('8.xlsx')

            x = x + 1

    # 最后关闭连接
    data1.close()
    data2.close()
    data3.close()
    data4.close()
    link.close()


if __name__ == "__main__":
    all_in()
