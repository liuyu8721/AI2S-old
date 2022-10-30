import datetime
import os
import re

import openpyxl as xl


def clean_blank_line(text):
    """去除文本中空行"""
    clear_lines = ""
    for line in text:
        if line != '\n':
            clear_lines += line
    return clear_lines


def q_to_b(uchar):
    """单个字符 全角转半角"""
    inside_code = ord(uchar)
    if inside_code == 12288:  # 全角空格直接转换
        inside_code = 32
    elif 65281 <= inside_code <= 65374:  # 全角字符（除空格）根据关系转化
        inside_code -= 65248
    return chr(inside_code)


def str_q_to_b(ustring):
    """把字符串全角转半角"""
    return "".join([q_to_b(uchar) for uchar in ustring])


def standardize_age(text):
    """将小于一岁的年龄标准化表示(年龄=月数/12)"""
    age_pattern = '(?:,)(\\d{1,2})个月(?:,)'
    match = re.search(age_pattern, text)
    if match:
        s = match.start()
        e = match.end()
        age = round(int(match.group(1)) / 12, 2)
        text = text[:s] + str(age) + '岁' + text[e:]
    return text


def clean_data(text):
    """处理替换文本中行动动词、标点符号、同义词,删除无用标点、字符"""

    def go_to(text_inner):
        """修改替换文本中不以句号结束的行动动词，统一为‘到’"""
        pattern = '(?:抵达|返回|进入|在|回|赴|回到|前往|去到|飞往)(?!\\。)'
        match = re.search(pattern, text_inner)
        while match:
            s = match.start()
            e = match.end()
            text_inner = text_inner[:s] + '到' + text_inner[e:]
            match = re.search(pattern, text_inner)
        return text_inner

    '''删除不表示时间的冒号,删除'--',删除空格'''
    text = re.sub(r'(?<=\D):|--|\s', '', text)
    '''删除序列标号 例如: 1. 3. 1、 2、'''
    text = re.sub(r'[1-9]\d*(?:\.(?:(?=\d月)|(?!\d))|、)', '', text)
    '''替换逗号,后括号 为 空格'''
    text = re.sub('[,)]', ' ', text)
    '''删除前括号'''
    text = re.sub('\\s?\\(', '', text)
    '''替换箭头(->或→)为句号'''
    text = re.sub(r'->|→', '。', text)
    '''替换分号为句号'''
    text = re.sub(r';', '。', text)
    '''删除无用干扰文字'''
    text = re.sub(r'微信关注椰城', '', text)
    '''将因前面处理后生成的句号而导致两个连续句号产生,替换为一个句号'''
    text = re.sub(r'。。', '。', text)
    '''修改替换文本中不以句号结束的行动动词，统一为‘到’'''
    text = go_to(text)
    '''替换文本中表示乘坐的动词，统一为‘乘’'''
    text = re.sub(r'乘坐|搭乘|转乘|坐', '乘', text)
    '''统一文本中火车的两种表示方式(动车或列车)为‘火车’'''
    text = re.sub(r'动车|列车', '火车', text)
    '''替换文本中‘飞机’为‘航班’'''
    text = re.sub(r'飞机', '航班', text)
    return text


def set_time(text):
    """修改替换文本中连续day的表示 例如: 2-7日 修改为 2日-7日"""
    pattern = '(\\d{1,2})-(\\d{1,2}日)'
    match = re.search(pattern, text)
    while match:
        s = match.start()
        e = match.end()
        day = match.group(1) + '日-' + match.group(2)
        text = text[:s] + day + text[e:]
        match = re.search(pattern, text)
    return text


def patch_month(text):
    """补全月份:将文本中没有月份表示的日期补全为带有月份的日期
    通过查找不含月份日期文本的前文或后文中含有月份表示的最近日期,比较日的大小,判断补全月份(月份不变或+1或-1)
    例如: 1月23日到琼山区龙塘镇仁三村委会道本村 25日上午11时从龙塘镇道本村到海口府城开始出车拉客 补全为:
    1月23日到琼山区龙塘镇仁三村委会道本村 1月25日上午11时从龙塘镇道本村到海口府城开始出车拉客"""
    pattern = '(?<!月|\\d)(\\d{1,2})日'
    match = re.search(pattern, text)
    while match:
        day = match.group(1)
        s = match.start()
        e = match.end()
        prev_text = text[:s]  # 不含月份的日期文本的前文
        next_text = text[e:]  # 不含月份的日期文本的后文
        p_time = re.findall('(\\d{1,2})月(\\d{1,2})日', prev_text)  # 查找前文是否含有带有月份的日期
        n_time = re.findall('(\\d{1,2})月(\\d{1,2})日', next_text)  # 查找后文是否含有带有月份的日期
        if p_time:
            '''查找到前文距离需补全日期最近的含月日期,比较日期'''
            if int(p_time[-1][1]) <= int(day):
                month = p_time[-1][0] + '月'  # 前文日期小于等于需补全日期，则需补全的月即为前文的月
            else:
                month = str(int(p_time[-1][0]) + 1) + '月'  # 前文日期大于需补全日期，则需补全的月即为前文的月+1
            date = month + day + '日'
            text = text[:s] + date + text[e:]
            match = re.search(pattern, text)
        elif n_time:
            '''查找到后文距离需补全日期最近的含月日期,比较日期'''
            if int(n_time[0][1]) < int(day):
                month = str(int(n_time[0][0]) - 1) + '月'  # 后文日期小于需补全日期，则需补全的月即为后文的月-1
            else:
                month = n_time[0][0] + '月'  # 后文日期大于等于需补全日期，则需补全的月即为后文的月
            date = month + day + '日'
            text = text[:s] + date + text[e:]
            match = re.search(pattern, text)
        else:
            break
    return text


def standardize_time(text):
    """
    标准化时间段表示:
    例如：
    1月24日中午-1月26日               →      1月24日中午~1月26日
    2019年12月18日-2020年1月19日上午  →      2019年12月18日~2020年1月19日上午
    1月25日晚-1月27日                 →       1月25日晚~1月27日
    2月13日晚至2月16日                →       2月13日晚~2月16日
    1月21-23日晚                     →       1月21日~1月23日晚
    1月20至24日                      →       1月20日~1月24日
    1月28日至30日                    →       1月28日~1月30日
    1月26日-2月1日                   →       1月26日~2月1日
    2020年1月25日下午至2月1日         →       2020年1月25日下午~2月1日
    """
    pattern = '(\\d{4}年)?(\\d{1,2})月(\\d{1,2})日?(上午|中午|下午|晚)?[-至](\\d{4}年)?(\\d{1,2})[日月](\\d{1,2})?日?(上午|中午|下午|晚)?'
    match = re.search(pattern, text)

    def ret_ut(var):
        """判断变量是否为None,不为None则返回变量,否则返回空字符"""
        if var:
            return var
        else:
            return ''

    while match:
        '''提取时间表示中可能含有的信息:'''
        s_year = ret_ut(match.group(1))  # 开始年份
        e_year = ret_ut(match.group(5))  # 结束年份
        s_time = ret_ut(match.group(4))  # 开始某日具体时间(上午|中午|晚)
        e_time = ret_ut(match.group(8))  # 结束某日具体时间(上午|中午|晚)
        s = match.start()
        e = match.end()
        if match.group(7):
            '''处理含有两个月份的时间段:例如:2月13日至2月16日'''
            # 'time = '$'+s_year+match.group(2)+'月'+match.group(3)+'日'+s_time+'~'+e_year+match.group(6)+'月'+match.group(7)+'日'+e_time+'$'
            time = s_year + match.group(2) + '月' + match.group(3) + '日' + s_time + '~' + e_year + \
                match.group(6) + '月' + match.group(7) + '日' + e_time
        else:
            '''处理只含有一个月份的时间段:例如:1月28日至30日'''
            # time = '$'+s_year+match.group(2)+'月'+match.group(3)+'日'+s_time+'~'+e_year+match.group(2)+'月'+match.group(6)+'日'+e_time+'$'
            time = s_year + match.group(2) + '月' + match.group(3) + '日' + s_time + '~' + e_year + \
                match.group(2) + '月' + match.group(6) + '日' + e_time
        text = text[:s] + time + text[e:]
        match = re.search(pattern, text)
    return text


def time_ch(text):
    """
    标准化时分表示:
    例如:
    13:30  →   13时30分
    16:00  →   16时
    """
    pattern = '(\\d{1,2}):(\\d{1,2})'
    match = re.search(pattern, text)
    while match:
        if match.group(2) == '00':
            hour = match.group(1)
            # time = '$'+hour+'时'+'$'
            time = hour + '时'
        else:
            hour = match.group(1)
            minute = match.group(2)
            # time = '$'+hour+'时'+minute+'分' + '$'
            time = hour + '时' + minute + '分'
        text = re.sub(pattern, time, text, count=1)
        match = re.search(pattern, text)
    text = re.sub('分分', '分', text)
    return text


def is_exist_train_and_flight(text):
    """寻找列车号和航班号,通过数据库查找是否存在"""
    train_pattern = '(?:车次|火车)\\s?([GCDZTSPKXLY1-9]\\d{1,4})|乘([GCDZTSPKXLY1-9]\\d{1,4})次火车'
    flight_pattern = '(?:国航|南航|海航|航班|航班号)\\s?([A-Z\\d]{2}\\d{3,4})|([A-Z\\d]{2}\\d{3,4})(?:次航班|航班)'

    def load_database(txts):
        """加载数据库"""
        f = open(txts, encoding='utf-8')
        txt = []
        for line in f:
            txt.append(line.strip())
        texts = ';'.join(txt)
        return texts

    # 加载列车号数据库
    train_number_databases = load_database('train_number.txt')
    # 加载航班号数据库
    flight_number_databases = load_database('flight.txt')

    # 寻找
    def find(type_inner, pattern, database, text_inner):
        """
        type_inner:编号类型(火车车次或航班)
        pattern:查找模型(火车模型或航班模型)
        database:加载数据库
        text_inner:需查找的文本
        """
        for match in re.finditer(pattern, text_inner):
            s = match.start()
            if text_inner[s - 1:s] == ' ':  # 删除航班或火车号前面的空格
                text_inner = text_inner[:s - 1] + text_inner[s:]
            if match.group(1) is not None:
                number = match.group(1)
            else:
                number = match.group(2)
            if re.search(number, database) is None:
                print(type_inner + number + ' not Exist')  # 查找数据库不存在的列车号和航班号
            # if re.search(number,database) != None:   查找数据库存在的列车号和航班号
            #     print(type_inner+number+' Exist')

    find('火车车次:', train_pattern, train_number_databases, text)
    find('航班号:', flight_pattern, flight_number_databases, text)
    return text


def modify_fuzzy_time(text):
    """
    模糊时间具体化:
    通过查找模糊时间词前面的日期,修改替换模糊时间词
    例如:
    2020年1月22日乘MU2527次航班从武汉到三亚→当晚乘私家车到三亚市天涯区水蛟村住处
    →
    2020年1月22日乘MU2527次航班从武汉到三亚→2020年1月15日晚上乘私家车到三亚市天涯区水蛟村住处
    """
    pattern = '当天|当晚'
    match = re.search(pattern, text)
    while match:
        time = re.findall('((?:\\d{4}年)?\\d{1,2}月\\d{1,2}日)(?=.+(?:当天|当晚))', text)
        s = match.start()
        e = match.end()
        if time:
            if match.group() == '当晚':
                # text = text[:s]+'$'+str(time[-1])+'晚上'+'$'+text[e:]
                text = text[:s] + str(time[-1]) + '晚上' + text[e:]
            else:
                text = text[:s] + str(time[-1]) + text[e:]
            # text = text[:s]+'$'+str(time[-1])+'$'+text[e:]
            match = re.search(pattern, text)
        else:
            break
    return text


def load_xlsx(file_path):
    """加载提取excel,对excel文本进行数据预处理"""
    lines_inner = []  # 列表存储文本数据
    wb = xl.load_workbook(file_path)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        uid = str(sheet.cell(row, 6).value) + '_' + \
              str(sheet.cell(row, 5).value) + '_' + \
              str(row - 1)  # 读取ID,并非从原文提取
        orig_text = sheet.cell(row, 22)  # 读取每段原文信息
        if orig_text.value is not None:  # 判断信息是否为空
            line = []
            # 去除文本中空行
            text = clean_blank_line(orig_text.value)
            # 存储去除空行后的原数据
            orig_text = text
            # 字符串全角转半角
            text = str_q_to_b(text)
            # 将小于一岁的年龄标准化表示
            text = standardize_age(text)
            # 处理替换文本中行动动词、标点符号、同义词,删除无用标点、字符
            text = clean_data(text)
            # 例如:2-7日 修改为 2日-7日
            text = set_time(text)
            # 补全月份
            text = patch_month(text)
            # 标准化时间段表示
            text = standardize_time(text)
            # text = modify_fuzzy_time(text)
            # 标准化时分表示
            text = time_ch(text)
            # 寻找列车号和航班号
            text = is_exist_train_and_flight(text)
            line.append(uid)
            line.append(orig_text)
            line.append(text)
            lines_inner.append(line)
    return lines_inner


def export_to_excel(path, lines_inner):
    """保存至excel文档"""
    if not os.path.exists(path):
        book = xl.Workbook()
        sheet = book.active
        sheet.append(('ID', 'originalTrack', 'Track'))
        book.save(path)

    wb = xl.load_workbook(path)
    sheet = wb['Sheet']
    r = 2
    for row in lines_inner:
        # sheet.append(row) #直接添加每行数据，问题：重复运行，会向下继续添加重复数据
        sheet.cell(r, column=1).value = row[0]  # 添加每行ID
        sheet.cell(r, column=2).value = row[1]  # 添加每行originalTrack
        sheet.cell(r, column=3).value = row[2]  # 添加每行Track
        r = r + 1
    wb.save(path)


def text_to_excel(file_path, path):
    lines_list = []
    lines_inner = ''
    separator = '****'
    file_read = open(file_path, 'r', encoding='UTF-8')
    line = file_read.readline().strip('\n')
    lines_inner = lines_inner + line + '。'
    while line:
        line = file_read.readline().strip('\n')
        if line == separator:
            lines_list.append(lines_inner)
            lines_inner = ''
            line = file_read.readline().strip('\n')
            lines_inner = lines_inner + line + '。'
        else:
            lines_inner = lines_inner + line + '。'
    file_read.close()

    if not os.path.exists(path):
        book = xl.Workbook()
        sheet = book.active
        titles = ['Track']
        sheet.append(titles)
        book.save(path)

    wb = xl.load_workbook(path)
    sheet = wb['Sheet']
    r = 2
    for row in lines_list:
        sheet.cell(r, column=3).value = row
        r = r + 1
    wb.save(path)


def split_text(file_path, bystop_path, ultim_path, excel_path):
    """
    按规则分隔文本:
    file_path:需分隔的excel文本路径
    bystop_path:通过句号为间隔符分隔处理后的text文本路径
    ultim_path:通过句号和日期分隔处理后的text文本路径
    path:通过分隔处理后的excel文本路径
    """
    # lines_inner = []  # 纯粹分隔处理后的文本
    split_by_stop(file_path, bystop_path)  # 以文本中句号分隔文本
    lines_inner = split_by_date(bystop_path)  # 以日期分隔文本
    lines_inner = clean_lines(lines_inner)  # 去除分隔后小于等于2个字符的段落
    # 文本存储为text
    file_write = open(ultim_path, 'w', encoding='UTF-8')
    for var in lines_inner:
        file_write.writelines(var)
        file_write.writelines('\n')
    file_write.close()
    # 文本覆盖纯粹为excel
    text_to_excel(ultim_path, excel_path)


def split_by_stop(excel_path, bystop_path):
    """
    以文本中句号分隔文本
    excel_path:需识别的excel文档地址
    bystop_path:分隔处理后生成的text文本地址
    """
    file_write = open(bystop_path, 'w', encoding='UTF-8')
    wb = xl.load_workbook(excel_path)  # 加载excel文本
    sheet = wb['Sheet']
    for row in range(2, sheet.max_row + 1):
        orig_text = str(sheet.cell(row, 3).value)
        # 添加ID:
        # ID =  str(sheet.cell(row,1).value)
        # file_write.writelines(ID)
        # file_write.writelines('\n')
        point = re.search('。', orig_text)
        while point:
            s = point.start()
            e = point.end()
            var = orig_text[:s]
            file_write.writelines(var)
            file_write.writelines('\n')
            orig_text = orig_text[e:]
            point = re.search('。', orig_text)
        file_write.writelines('****\n')
    file_write.close()


def split_by_date(file_path):
    """以日期分隔文本"""
    # lines_list = []
    lines_list = simple_split_by_date(file_path)
    lines_list = more_split_by_date(lines_list)
    # CDLines = split_by_ctn_date(DLines)
    lines_list = get_to_date(lines_list)
    return lines_list


def simple_split_by_date(file_path):
    """
    识别每行，判断是否含有日期，含有则不做处理，不含有则将改行添加到前一行
    """
    separator = '****'  # 分隔符
    lines_inner = []  # 读取存储文本每行
    # 读取经过句号分隔的文本
    file_read = open(file_path, 'r', encoding='UTF-8')
    line = file_read.readline().strip('\n')
    lines_inner.append(line)
    # line = file_read.readline().strip('\n')   #添加ID时需要
    # lines_inner.append(line)
    while line:
        line = file_read.readline().strip('\n')  # 去除换行符
        if line == separator:  # 判断是否为分隔标记
            lines_inner.append(line)
            line = file_read.readline().strip('\n')
            lines_inner.append(line)
            # line = file_read.readline().strip('\n')  #添加ID时需要
            # lines_inner.append(line)
            continue
        date_pattern = '\\d{1,2}月(?:\\d{1,2}日)?'
        date_match = re.search(date_pattern, line)
        if date_match is None:
            lines_inner[-1] = lines_inner[-1] + ' ' + line
        else:
            lines_inner.append(line)
    file_read.close()
    return lines_inner


def more_split_by_date(lines_inner):
    """
    若为0个，直接添加；
    若为1个，判断该日期是否在行首出现，
        若在，则直接添加，
        不在，则通过日期出现的首位置分隔文本，判断前一行是否为间隔符，
            不是则将日期前文本添加到前一行，日期后文本成为独立一行
            是则将日期前文本独立成一行，日期后文本成为独立一行；
    若为2个及以上，通过判断每个日期出现位置，分隔成多行。
    """
    separator = '****'  # 分隔符
    d_lines = []
    for line in lines_inner:
        date_pattern = '(?:(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?(?:~(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?)?[和及、])+(?:(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?(?:~(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?)?)|(?:(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?~(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?)|(?:(?<!\\~)(?:\\d{4}年)?\\d{1,2}月(?:\\d{1,2}日)?(?:上午|中午|晚)?(?!\\~))'
        # SeveralDatePattern = '(?:(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?(?:~(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?)?[和及、])+(?:(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?(?:~(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?)?)'
        date_s = len(re.findall(date_pattern, line))
        if date_s == 0:
            d_lines.append(line)
        elif date_s == 1:
            pattern = '^(?:\\d{4}年)?\\d{1,2}月\\d{1,2}日(?:上午|中午|晚)?'
            if re.search(pattern, line):  # 查看是否在行首出现月日
                d_lines.append(line)
            else:
                match = re.search(date_pattern, line)
                s = match.start()
                if d_lines[-1] == separator:  # 判断上一句是否为间隔符
                    d_lines.append(line[:s])
                else:
                    d_lines[-1] = d_lines[-1] + ' ' + line[:s]  # s添加到上一句末尾
                d_lines.append(line[s:])
        else:
            # print(line)
            date_match = re.finditer(date_pattern, line)
            last_date_sep = 0
            flag = True  # 文本以日期开头，则Flag为False
            for date in date_match:
                date_sep = date.start()
                if date_sep != 0:
                    if last_date_sep == 0 and flag:
                        # if checkdata(d_lines[-1]):  #检查上一句是否有日期
                        if d_lines[-1] == separator:  # 判断上一句是否为间隔符
                            d_lines.append(line[last_date_sep:date_sep])
                        else:
                            d_lines[-1] = d_lines[-1] + ' ' + line[last_date_sep:date_sep]  # 添加到上一句末尾
                        last_date_sep = date_sep
                    else:
                        d_lines.append(line[last_date_sep:date_sep])
                        last_date_sep = date_sep
                        flag = True
                else:
                    flag = False
            d_lines.append(line[last_date_sep:])
    return d_lines


def get_to_date(lines_inner):
    """
    通过前面的处理，会将日期在句子后面格式的文本错误分隔，例如：
    1月26日早上体感不适 居家休息至1月27日
    分隔成
    1月26日早上体感不适 居家休息至\n1月27日
    因此，我们可以通过识别每行，判断是否只含有日期，将其重新修正为：
    1月26日早上体感不适\n居家休息至1月27日
    """
    o_lines = []
    pattern = '^(?:(\\d{4})年)?\\d{1,2}月\\d{1,2}日$'
    pattern2 = '(?<=\\s)(\\S+至)$'
    for var in lines_inner:
        match = re.search(pattern, var)
        if match:
            match2 = re.search(pattern2, o_lines[-1])
            if match2:
                text = match2.group(1)
                line = text + match.group()
                o_lines[-1] = o_lines[-1][:match2.start()]
                o_lines.append(line)
        else:
            o_lines.append(var)
    return o_lines


def split_by_ctn_date(lines_inner):
    """
    连续段日期处理(可选)
    将连续日期，例如：
    1月24日~1月26日到老家过年
    分隔为：
    1月24日到老家过年
    1月25日到老家过年
    1月26日到老家过年
    """
    cd_lines = []
    for var in lines_inner:
        pattern = '(?:(\\d{4})年)?(\\d{1,2})月(\\d{1,2})日\\~(\\d{1,2})月(\\d{1,2})日'
        match = re.search(pattern, var)
        if match:
            e = match.end()
            start_month = int(match.group(2))
            start_day = int(match.group(3))
            end_month = int(match.group(4))

            end_day = int(match.group(5))
            year = 1970
            if match.group(1) is not None:
                year = int(match.group(1))

            text = var[e:]
            begin = datetime.date(year, start_month, start_day)
            end = datetime.date(year, end_month, end_day)
            delta = datetime.timedelta(days=1)
            if match.group(1) is not None:
                while begin <= end:
                    c_text = begin.strftime("%Y年%m月%d日") + text
                    cd_lines.append(c_text)
                    begin += delta
            else:
                while begin <= end:
                    c_text = begin.strftime("%m月%d日") + text
                    cd_lines.append(c_text)
                    begin += delta
        else:
            cd_lines.append(var)
    return cd_lines


def clean_lines(lines_inner):
    """去除分隔后小于等于2个字符的段落"""
    cleaned = []
    for line in lines_inner:
        if len(line) > 2:
            cleaned.append(line)
    return cleaned


if __name__ == '__main__':
    lines = load_xlsx("CRF_Hainan.xlsx")
    export_to_excel("hainan_Track15.xlsx", lines)
    split_text("hainan_Track15.xlsx", "SPlitByStop.txt", "SplitTextUltimate.txt", "hainan_Track15.xlsx")
