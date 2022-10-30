from __future__ import division
import openpyxl
import sqlite3
from pymysql import *
import re
import os
import time
from LAC import LAC
import openpyxl as xl
import datetime
import xlrd
import random
import requests
from xlwt import *
from urllib.parse import *


def clearBlankLine(text):
    """去除文本中空行"""
    clear_lines = ""
    for line in text:
        if line != '\n':
            clear_lines += line
    return clear_lines


def Q2B(uchar):
    """单个字符 全角转半角"""
    inside_code = ord(uchar)
    if inside_code == 12288:  # 全角空格直接转换
        inside_code = 32
    elif (inside_code >= 65281 and inside_code <= 65374):  # 全角字符（除空格）根据关系转化
        inside_code -= 65248
    return chr(inside_code)


def stringQ2B(ustring):
    """把字符串全角转半角"""
    return "".join([Q2B(uchar) for uchar in ustring])


def StandardAge(text):
    """将小于一岁的年龄标准化表示(年龄=月数/12)"""
    AgePattern = '(?:,)(\d{1,2})个月(?:,)'
    match = re.search(AgePattern, text)
    if match:
        s = match.start()
        e = match.end()
        age = round(int(match.group(1)) / 12, 2)
        text = text[:s] + str(age) + '岁' + text[e:]
    return text


def clean_data(text):
    """处理替换文本中行动动词、标点符号、同义词,删除无用标点、字符"""

    def goTo(text):
        '''修改替换文本中不以句号结束的行动动词，统一为‘到’'''
        Pattern = '(?:抵达|返回|进入|在|回|赴|回到|前往|去到|飞往)(?!\。)'
        Match = re.search(Pattern, text)
        while Match:
            s = Match.start()
            e = Match.end()
            text = text[:s] + '到' + text[e:]
            Match = re.search(Pattern, text)
        return text

    '''删除不表示时间的冒号,删除'--',删除空格'''
    text = re.sub(r'(?<=\D):|--|\s', '', text)
    '''删除序列标号 例如: 1. 3. 1、 2、'''
    text = re.sub(r'[1-9]\d*(?:\.(?:(?=\d月)|(?!\d))|\、)', '', text)
    '''替换逗号,后括号 为 空格'''
    text = re.sub('[,)]', ' ', text)
    '''删除前括号'''
    text = re.sub('\s?\(', '', text)
    '''替换箭头(->或→)为句号'''
    text = re.sub(r'->|→', '。', text)
    '''替换分号为句号'''
    text = re.sub(r';', '。', text)
    '''删除无用干扰文字'''
    text = re.sub(r'微信关注椰城', '', text)
    '''将因前面处理后生成的句号而导致两个连续句号产生,替换为一个句号'''
    text = re.sub(r'。。', '。', text)
    '''修改替换文本中不以句号结束的行动动词，统一为‘到’'''
    text = goTo(text)
    '''替换文本中表示乘坐的动词，统一为‘乘’'''
    text = re.sub(r'乘坐|搭乘|转乘|坐', '乘', text)
    '''统一文本中火车的两种表示方式(动车或列车)为‘火车’'''
    text = re.sub(r'动车|列车', '火车', text)
    '''替换文本中‘飞机’为‘航班’'''
    text = re.sub(r'飞机', '航班', text)
    return text


def settime(text):
    '''修改替换文本中连续day的表示 例如: 2-7日 修改为 2日-7日'''
    Pattern = '(\d{1,2})-(\d{1,2}日)'
    match = re.search(Pattern, text)
    while match:
        s = match.start()
        e = match.end()
        day = match.group(1) + '日-' + match.group(2)
        text = text[:s] + day + text[e:]
        match = re.search(Pattern, text)
    return text


def PatchMonth(text):
    '''补全月份:将文本中没有月份表示的日期补全为带有月份的日期
    通过查找不含月份日期文本的前文或后文中含有月份表示的最近日期,比较日的大小,判断补全月份(月份不变或+1或-1)
    例如: 1月23日到琼山区龙塘镇仁三村委会道本村 25日上午11时从龙塘镇道本村到海口府城开始出车拉客 补全为:
    1月23日到琼山区龙塘镇仁三村委会道本村 1月25日上午11时从龙塘镇道本村到海口府城开始出车拉客'''
    Pattern = '(?<!月|\d)(\d{1,2})日'
    match = re.search(Pattern, text)
    while match:
        day = match.group(1)
        s = match.start()
        e = match.end()
        prevtext = text[:s]  # 不含月份的日期文本的前文
        nexttext = text[e:]  # 不含月份的日期文本的后文
        ptime = re.findall('(\d{1,2})月(\d{1,2})日', prevtext)  # 查找前文是否含有带有月份的日期
        ntime = re.findall('(\d{1,2})月(\d{1,2})日', nexttext)  # 查找后文是否含有带有月份的日期
        if ptime:
            '''查找到前文距离需补全日期最近的含月日期,比较日期'''
            if int(ptime[-1][1]) <= int(day):
                month = ptime[-1][0] + '月'  # 前文日期小于等于需补全日期，则需补全的月即为前文的月
            else:
                month = str(int(ptime[-1][0]) + 1) + '月'  # 前文日期大于需补全日期，则需补全的月即为前文的月+1
            date = month + day + '日'
            text = text[:s] + date + text[e:]
            match = re.search(Pattern, text)
        elif ntime:
            '''查找到后文距离需补全日期最近的含月日期,比较日期'''
            if int(ntime[0][1]) < int(day):
                month = str(int(ntime[0][0]) - 1) + '月'  # 后文日期小于需补全日期，则需补全的月即为后文的月-1
            else:
                month = ntime[0][0] + '月'  # 后文日期大于等于需补全日期，则需补全的月即为后文的月
            date = month + day + '日'
            text = text[:s] + date + text[e:]
            match = re.search(Pattern, text)
        else:
            break
    return text


def StandardTime(text):
    '''
    标准化时间段表示:
    例如：
    1月24日中午-1月26日               →      1月24日中午~1月26日
    2019年12月18日-2020年1月19日上午  →      2019年12月18日~2020年1月19日上午
    1月25日晚-1月27日                 →       1月25日晚~1月27日
    2月13日晚至2月16日                →       2月13日晚~2月16日
    1月21-23日晚                     →       1月21日~1月23日晚
    1月20至24日                      →       1月20日~1月24日
    1月28日至30日                    →       1月28日~1月30日
    1月26日-2月1日                   →       1月26日~2月1日
    2020年1月25日下午至2月1日         →       2020年1月25日下午~2月1日
    '''
    Pattern = '(\d{4}年)?(\d{1,2})月(\d{1,2})日?(上午|中午|下午|晚)?[-至](\d{4}年)?(\d{1,2})[日月](\d{1,2})?日?(上午|中午|下午|晚)?'
    match = re.search(Pattern, text)

    def retut(var):
        '''判断变量是否为None,不为None则返回变量,否则返回空字符'''
        if var:
            return var
        else:
            return ''

    while match:
        '''提取时间表示中可能含有的信息:'''
        SYear = retut(match.group(1))  # 开始年份
        EYear = retut(match.group(5))  # 结束年份
        Stime = retut(match.group(4))  # 开始某日具体时间(上午|中午|晚)
        Etime = retut(match.group(8))  ##结束某日具体时间(上午|中午|晚)
        s = match.start()
        e = match.end()
        if match.group(7):
            '''处理含有两个月份的时间段:例如:2月13日至2月16日'''
            # 'time = '$'+SYear+match.group(2)+'月'+match.group(3)+'日'+Stime+'~'+EYear+match.group(6)+'月'+match.group(7)+'日'+Etime+'$'
            time = SYear + match.group(2) + '月' + match.group(3) + '日' + Stime + '~' + EYear + match.group(
                6) + '月' + match.group(7) + '日' + Etime
        else:
            '''处理只含有一个月份的时间段:例如:1月28日至30日'''
            # time = '$'+SYear+match.group(2)+'月'+match.group(3)+'日'+Stime+'~'+EYear+match.group(2)+'月'+match.group(6)+'日'+Etime+'$'
            time = SYear + match.group(2) + '月' + match.group(3) + '日' + Stime + '~' + EYear + match.group(
                2) + '月' + match.group(6) + '日' + Etime
        text = text[:s] + time + text[e:]
        match = re.search(Pattern, text)
    return text


def timech(text):
    '''
    标准化时分表示:
    例如:
    13:30  →   13时30分
    16:00  →   16时
    '''
    Pattern = '(\d{1,2}):(\d{1,2})'
    match = re.search(Pattern, text)
    while match:
        if match.group(2) == '00':
            hour = match.group(1)
            # time = '$'+hour+'时'+'$'
            time = hour + '时'
        else:
            hour = match.group(1)
            minute = match.group(2)
            # time = '$'+hour+'时'+minute+'分' + '$'
            time = hour + '时' + minute + '分'
        text = re.sub(Pattern, time, text, count=1)
        match = re.search(Pattern, text)
    text = re.sub('分分', '分', text)
    return text


def IsExistTrainAndFlight(text):
    '''寻找列车号和航班号,通过数据库查找是否存在'''
    TrainPattern = '(?:车次|火车)\s?([GCDZTSPKXLY1-9]\d{1,4})|乘([GCDZTSPKXLY1-9]\d{1,4})次火车'
    FlightPattern = '(?:国航|南航|海航|航班|航班号)\s?([A-Z\d]{2}\d{3,4})|([A-Z\d]{2}\d{3,4})(?:次航班|航班)'

    def LoadDataBase(txts):
        '''加载数据库'''
        f = open(txts, encoding='utf-8')
        txt = []
        for line in f:
            txt.append(line.strip())
        texts = ';'.join(txt)
        return texts

    # 加载列车号数据库
    TrainNumberDataBases = LoadDataBase('train_number.txt')
    # 加载航班号数据库
    FlightNumberDataBases = LoadDataBase('flight.txt')

    # 寻找
    def find(type, Pattern, DataBase, text):
        '''
        type:编号类型(火车车次或航班)
        Pattern:查找模型(火车模型或航班模型)
        DataBase:加载数据库
        text:需查找的文本
        '''
        for Match in re.finditer(Pattern, text):
            s = Match.start()
            if text[s - 1:s] == ' ':  # 删除航班或火车号前面的空格
                text = text[:s - 1] + text[s:]
            if Match.group(1) != None:
                Number = Match.group(1)
            else:
                Number = Match.group(2)
            if re.search(Number, DataBase) == None:
                print(type + Number + ' not Exist')  # 查找数据库不存在的列车号和航班号

    find('火车车次:', TrainPattern, TrainNumberDataBases, text)
    find('航班号:', FlightPattern, FlightNumberDataBases, text)
    return text


def ModifyBlurTime(text):
    '''
    模糊时间具体化:
    通过查找模糊时间词前面的日期,修改替换模糊时间词
    例如:
    2020年1月22日乘MU2527次航班从武汉到三亚→当晚乘私家车到三亚市天涯区水蛟村住处
    →
    2020年1月22日乘MU2527次航班从武汉到三亚→2020年1月15日晚上乘私家车到三亚市天涯区水蛟村住处
    '''
    Pattern = '当天|当晚'
    match = re.search(Pattern, text)
    while match:
        time = re.findall('((?:\d{4}年)?\d{1,2}月\d{1,2}日)(?=.+(?:当天|当晚))', text)
        s = match.start()
        e = match.end()
        if time:
            if match.group() == '当晚':
                # text = text[:s]+'$'+str(time[-1])+'晚上'+'$'+text[e:]
                text = text[:s] + str(time[-1]) + '晚上' + text[e:]
            else:
                text = text[:s] + str(time[-1]) + text[e:]
            # text = text[:s]+'$'+str(time[-1])+'$'+text[e:]
            match = re.search(Pattern, text)
        else:
            break
    return text


def load_xlsx(file_path):
    '''加载提取excel,对excel文本进行数据预处理'''
    lines = []  # 列表存储文本数据
    wb = xl.load_workbook(file_path)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        ID = str(sheet.cell(row, 6).value) + '_' + str(sheet.cell(row, 5).value) + '_' + str(row - 1)  # 读取ID,并非从原文提取
        orig_text = sheet.cell(row, 22)  # 读取每段原文信息
        if orig_text.value != None:  # 判断信息是否为空
            line = []
            # 去除文本中空行
            text = clearBlankLine(orig_text.value)
            # 存储去除空行后的原数据
            orig_text = text
            # 字符串全角转半角
            text = stringQ2B(text)
            # 将小于一岁的年龄标准化表示
            text = StandardAge(text)
            # 处理替换文本中行动动词、标点符号、同义词,删除无用标点、字符
            text = clean_data(text)
            # 例如:2-7日 修改为 2日-7日
            text = settime(text)
            # 补全月份
            text = PatchMonth(text)
            # 标准化时间段表示
            text = StandardTime(text)
            # text = ModifyBlurTime(text)
            # 标准化时分表示
            text = timech(text)
            # 寻找列车号和航班号
            text = IsExistTrainAndFlight(text)
            line.append(ID)
            line.append(orig_text)
            line.append(text)
            lines.append(line)
    return lines


def Save_to_excle(Save_Path_textToexcel, lines):
    '''保存至excel文档'''
    if not os.path.exists(Save_Path_textToexcel):
        book = xl.Workbook()
        sheet = book.active
        sheet.append(('ID', 'originalTrack', 'Track'))
        book.save(Save_Path_textToexcel)

    wb = xl.load_workbook(Save_Path_textToexcel)
    sheet = wb['Sheet']
    r = 2
    for row in lines:
        # sheet.append(row) #直接添加每行数据，问题：重复运行，会向下继续添加重复数据
        sheet.cell(r, column=1).value = row[0]  # 添加每行ID
        sheet.cell(r, column=2).value = row[1]  # 添加每行originalTrack
        sheet.cell(r, column=3).value = row[2]  # 添加每行Track
        r = r + 1
    wb.save(Save_Path_textToexcel)


def textToexcel(file_path, Save_Path_textToexcel):
    LinesList = []
    Lines = ''
    separator = '****'
    file_read = open(file_path, 'r', encoding='UTF-8')
    line = file_read.readline().strip('\n')
    Lines = Lines + line + '。'
    while line:
        line = file_read.readline().strip('\n')
        if line == separator:
            LinesList.append(Lines)
            Lines = ''
            line = file_read.readline().strip('\n')
            Lines = Lines + line + '。'
        else:
            Lines = Lines + line + '。'
    file_read.close()

    if not os.path.exists(Save_Path_textToexcel):
        book = xl.Workbook()
        sheet = book.active
        titles = []
        titles.append('Track')
        sheet.append(titles)
        book.save(Save_Path_textToexcel)

    wb = xl.load_workbook(Save_Path_textToexcel)
    sheet = wb['Sheet']
    r = 2
    for row in LinesList:
        sheet.cell(r, column=3).value = row
        r = r + 1
    wb.save(Save_Path_textToexcel)


def SplitText(file_path, Save_path_ByStop, Save_Path_ultim, Save_Path_textToexcel):
    '''
    按规则分隔文本:
    file_path:需分隔的excel文本路径
    Save_path_ByStop:通过句号为间隔符分隔处理后的text文本路径
    Save_Path_ultim:通过句号和日期分隔处理后的text文本路径
    Save_Path_textToexcel:通过分隔处理后的excel文本路径
    '''
    LInes = []  # 纯粹分隔处理后的文本
    SplitByStop(file_path, Save_path_ByStop)  # 以文本中句号分隔文本
    LInes = SplitByDate(Save_path_ByStop)  # 以日期分隔文本
    LInes = ClearLines(LInes)  # 去除分隔后小于等于2个字符的段落
    # 文本存储为text
    file_write = open(Save_Path_ultim, 'w', encoding='UTF-8')
    for var in LInes:
        file_write.writelines(var)
        file_write.writelines('\n')
    file_write.close()
    # 文本覆盖纯粹为excel
    textToexcel(Save_Path_ultim, Save_Path_textToexcel)


def SplitByStop(excelfile_path, Save_textpath_ByStop):
    '''
    以文本中句号分隔文本
    excelfile_path:需识别的excel文档地址
    Save_textpath_ByStop:分隔处理后生成的text文本地址
    '''
    file_write = open(Save_textpath_ByStop, 'w', encoding='UTF-8')
    wb = xl.load_workbook(excelfile_path)  # 加载excel文本
    sheet = wb['Sheet']
    for row in range(2, sheet.max_row + 1):
        orig_text = str(sheet.cell(row, 3).value)
        # 添加ID:
        # ID =  str(sheet.cell(row,1).value)
        # file_write.writelines(ID)
        # file_write.writelines('\n')
        point = re.search('。', orig_text)
        while point:
            s = point.start()
            e = point.end()
            var = orig_text[:s]
            file_write.writelines(var)
            file_write.writelines('\n')
            orig_text = orig_text[e:]
            point = re.search('。', orig_text)
        file_write.writelines('****\n')
    file_write.close()


def SplitByDate(file_path):
    '''以日期分隔文本'''
    LinesList = []
    LinesList = SimpleSplitByDate(file_path)
    LinesList = MoreSplitByDate(LinesList)
    # CDLines = SplitByCTNDate(DLines)
    LinesList = GetToDate(LinesList)
    return LinesList


def SimpleSplitByDate(file_path):
    '''
    识别每行，判断是否含有日期，含有则不做处理，不含有则将改行添加到前一行
    '''
    separator = '****'  # 分隔符
    Lines = []  # 读取存储文本每行
    # 读取经过句号分隔的文本
    file_read = open(file_path, 'r', encoding='UTF-8')
    line = file_read.readline().strip('\n')
    Lines.append(line)
    # line = file_read.readline().strip('\n')   #添加ID时需要
    # Lines.append(line)
    while line:
        line = file_read.readline().strip('\n')  # 去除换行符
        if line == separator:  # 判断是否为分隔标记
            Lines.append(line)
            line = file_read.readline().strip('\n')
            Lines.append(line)
            # line = file_read.readline().strip('\n')  #添加ID时需要
            # Lines.append(line)
            continue
        datePattern = '\d{1,2}月(?:\d{1,2}日)?'
        DateMatch = re.search(datePattern, line)
        if DateMatch == None:
            Lines[-1] = Lines[-1] + ' ' + line
        else:
            Lines.append(line)
    file_read.close()
    return Lines


def MoreSplitByDate(Lines):
    '''
    若为0个，直接添加；
    若为1个，判断该日期是否在行首出现，
        若在，则直接添加，
        不在，则通过日期出现的首位置分隔文本，判断前一行是否为间隔符，
            不是则将日期前文本添加到前一行，日期后文本成为独立一行
            是则将日期前文本独立成一行，日期后文本成为独立一行；
    若为2个及以上，通过判断每个日期出现位置，分隔成多行。
    '''
    separator = '****'  # 分隔符
    DLines = []
    for line in Lines:
        DatePattern = '(?:(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?(?:~(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?)?[和及、])+(?:(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?(?:~(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?)?)|(?:(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?~(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?)|(?:(?<!\~)(?:\d{4}年)?\d{1,2}月(?:\d{1,2}日)?(?:上午|中午|晚)?(?!\~))'
        # SeveralDatePattern = '(?:(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?(?:~(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?)?[和及、])+(?:(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?(?:~(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?)?)'
        dateS = len(re.findall(DatePattern, line))
        if dateS == 0:
            DLines.append(line)
        elif dateS == 1:
            Pattern = '^(?:\d{4}年)?\d{1,2}月\d{1,2}日(?:上午|中午|晚)?'
            if re.search(Pattern, line):  # 查看是否在行首出现月日
                DLines.append(line)
            else:
                Match = re.search(DatePattern, line)
                s = Match.start()
                if DLines[-1] == separator:  # 判断上一句是否为间隔符
                    DLines.append(line[:s])
                else:
                    DLines[-1] = DLines[-1] + ' ' + line[:s]  ##添加到上一句末尾
                DLines.append(line[s:])
        else:
            DateMatch = re.finditer(DatePattern, line)
            lastdatesep = 0
            Flag = True  # 文本以日期开头，则Flag为False
            for date in DateMatch:
                datesep = date.start()
                if datesep != 0:
                    if lastdatesep == 0 and Flag:
                        # if checkdata(DLines[-1]):  #检查上一句是否有日期
                        if DLines[-1] == separator:  # 判断上一句是否为间隔符
                            DLines.append(line[lastdatesep:datesep])
                        else:
                            DLines[-1] = DLines[-1] + ' ' + line[lastdatesep:datesep]  # 添加到上一句末尾
                        lastdatesep = datesep
                    else:
                        DLines.append(line[lastdatesep:datesep])
                        lastdatesep = datesep
                        Flag = True
                else:
                    Flag = False
            DLines.append(line[lastdatesep:])
    return DLines


def GetToDate(Lines):
    '''
    通过前面的处理，会将日期在句子后面格式的文本错误分隔，例如：
    1月26日早上体感不适 居家休息至1月27日
    分隔成
    1月26日早上体感不适 居家休息至\n1月27日
    因此，我们可以通过识别每行，判断是否只含有日期，将其重新修正为：
    1月26日早上体感不适\n居家休息至1月27日
    '''
    OLines = []
    Pattern = '^(?:(\d{4})年)?\d{1,2}月\d{1,2}日$'
    Pattern2 = '(?<=\s)(\S+至)$'
    for var in Lines:
        Match = re.search(Pattern, var)
        if Match:
            Match2 = re.search(Pattern2, OLines[-1])
            if Match2:
                text = Match2.group(1)
                line = text + Match.group()
                OLines[-1] = OLines[-1][:Match2.start()]
                OLines.append(line)
        else:
            OLines.append(var)
    return OLines


def SplitByCTNDate(Lines):
    '''
    连续段日期处理(可选)
    将连续日期，例如：
    1月24日~1月26日到老家过年
    分隔为：
    1月24日到老家过年
    1月25日到老家过年
    1月26日到老家过年
    '''
    CDLines = []
    for var in Lines:
        Pattern = '(?:(\d{4})年)?(\d{1,2})月(\d{1,2})日\~(\d{1,2})月(\d{1,2})日'
        Match = re.search(Pattern, var)
        if Match:
            e = Match.end()
            startMonth = int(Match.group(2))
            startDay = int(Match.group(3))
            endMonth = int(Match.group(4))

            endDay = int(Match.group(5))
            year = 1970
            if Match.group(1) != None:
                year = int(Match.group(1))

            text = var[e:]
            begin = datetime.date(year, startMonth, startDay)
            end = datetime.date(year, endMonth, endDay)
            delta = datetime.timedelta(days=1)
            if Match.group(1) != None:
                while begin <= end:
                    Ctext = begin.strftime("%Y年%m月%d日") + text
                    CDLines.append(Ctext)
                    begin += delta
            else:
                while begin <= end:
                    Ctext = begin.strftime("%m月%d日") + text
                    CDLines.append(Ctext)
                    begin += delta
        else:
            CDLines.append(var)
    return CDLines


def ClearLines(Lines):
    '''去除分隔后小于等于2个字符的段落'''
    ClearLines = []
    for line in Lines:
        if len(line) > 2:
            ClearLines.append(line)
    return ClearLines


def original_data(original_text):
    """
    :return: 返回的是经过对不规则符号进行处理之后的数据
    """
    return [original_text]


def extend_data(file_path, id_list):
    """_summary_

    Args:
        file_path (_type_): _description_
        id (_type_): _description_

    Returns:
        _type_: _description_
    """
    s_number = []
    s_info = []
    s_number_aligned = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    for i in range(2, sheet.max_row + 1):
        s_number.append(sheet['F' + str(i)].value)
        s_info.append(sheet['V' + str(i)].value)
    s_number = [line for line in s_number if line is not None]
    s_info = [line for line in s_info if line is not None]

    # 清洗
    s_number = [re.sub("[\=,\"]", repl="", string=line) for line in s_number]
    s_info = [re.match("第[0-9]+号", line).group(0)[:-1] for line in s_info]

    print("s_number:", s_number)
    print("s_info:", s_info)

    # 和id_list中标签对齐
    for i in range(len(id_list)):
        if id_list[i] in s_info:
            s_number_aligned.append(s_number[s_info.index(id_list[i])])
        else:
            s_number_aligned.append("未知")
    print("s_number_aligned:", s_number_aligned)

    return s_number_aligned


def read_res_excel(file_path):
    s = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    # 收集确诊病例号
    for i in range(2, sheet.max_row + 1):
        s.append(sheet['B' + str(i)].value)
    s = [line for line in s if line is not None]
    return wb, sheet, s


def merge(extend_file, res_file):
    """
    Args:
        extend_file (_type_): 额外excel文件，比如这里的excel.xlsx
        res_file (_type_): 生成的结果文件,比如海南疫情.xlsx
    """
    # step1.读取结果文件
    wb_res, sheet_res, s_res = read_res_excel(file_path=res_file)
    print("s_res:\n", s_res, len(s_res))

    # step2.读取额外文件
    s_number_aligned = extend_data(file_path=extend_file, id_list=s_res)
    # step3.插入到第一行

    sheet_res.insert_cols(1, 1)  # idx:插入的位置，用数字3，amount:插入的列数
    sheet_res['A' + str(1)].value = '唯一编号'
    for i in range(0, len(s_res)):
        sheet_res['A' + str(i + 2)].value = s_number_aligned[i]
    # step4 写入
    wb_res.save('./merged.xlsx')


def data_handle(text):
    sheet = []
    sheet.append(
        ('原文', '病例编号', '性别', '年龄', '常住地', '关系', '日期', '句子', '交通编号', '交通方式', '出发点', '目的地'))

    s3 = []  # 将数据用句号进行切割
    s4 = []  # 将数据用句号进行切割的数据
    list1 = []
    s1 = original_data(text)
    data_list = []  # 将每个文本的信息存储到列表中
    for i in s1:  # 从对不规则符号经过处理之后的数据中取值，每次取出的是excel表中每一行存储的文本数据
        flag = False
        routes = []
        nums = re.search(r'第\d+', i)
        if nums:
            num = nums.group()
        else:
            num = ''
        sex = re.search(r'[男|女]', i)  # 正则得到性别
        ages = re.search(r'\d+[\u5c81]', i)  # 正则得到年龄
        citys = re.search(r'常住\w+', i)
        if citys:
            city = citys.group()
        else:
            city = ''
        relation1 = re.search(r'(\d+号\、?\w+病例朋友)', i)
        relation2 = re.search(r'(系\d[\w+\、?]+)|(与第\w+[\、?\w+]+)(亲属关系|亲友关系|同一小区|一起乘网约车)', i)
        if relation2 and relation1:
            s = relation2.group() + ',' + relation1.group()
            flag = True
        if relation2:
            if flag:
                relation = s
            else:
                relation = relation2.group()
        else:
            relation = ''
        s3 = (i.split('。'))  # 文本中的数据做句号切割，因为每个文本中包含多句话，而每句话中包含多个需要的信息
        for j in s3:  # 取出来的是每句话
            if j == '\n' or j == '':
                continue
            else:
                time = re.findall(r'\d+月\d+日', j)  # 正则得到时间
                if time != []:
                    for t in time:
                        list1 = []
                        dict_y = []
                        sentence = j.split(t)[1]
                        times = re.search(r'\d+月\d+日', sentence)
                        sentences = re.findall(
                            r'(乘\w+|转运至\w+|去\w+|到\w+|到达\w+|从\w+|在\w+|进入\w+|返回\w+|抵达\w+|转至\w+|入住\w+|前往\w+|送至\w+|接往\w+|转至\w+|乘坐\w+)',
                            sentence)
                        if times and len(sentences) > 1:
                            list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence.split('，')[0]])
                            if sentence != []:
                                sheet.append(process(sentence, list1))
                        elif 1 < len(sentences):
                            for z in range(len(sentences)):
                                list1 = []
                                list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence])
                                if sentence[z] != []:
                                    rest = process(sentences[z], list1).copy()
                                    sheet.append(rest)
                        else:
                            list1.extend([i, num, sex.group(), ages.group(), city, relation, t, sentence])
                            if sentence != []:
                                sheet.append(process(sentence, list1))
    return sheet


def process(lines, list1):
    """
    Args:
        in_text (_type_): _description_
        out_text (_type_): _description_
    """
    # 装载LAC模型
    lac = LAC(mode='lac')
    s = re.search(r'机场', lines)
    s1 = re.search(r'市妇幼保健院', lines)
    s2 = re.search(r'琼海市博鳌镇幸福海小区家中', lines)
    s3 = re.search(r'市人民医院', lines)
    s4 = re.search(r'大小洞天', lines)
    s5 = re.search('r天涯区升升超市', lines)
    s6 = re.search(r'河西路口好芙利蛋糕店', lines)
    s7 = re.search(r'川味饭店', lines)
    s8 = re.search(r'海鲜广场', lines)
    s9 = re.search(r'一心堂药店', lines)
    s10 = re.search(r'群众街', lines)
    s11 = re.search(r'川味王餐厅', lines)
    s12 = re.search(r'羊栏市场', lines)
    s13 = re.search(r'县人民医院', lines)
    s14 = re.search(r'省人民医院', lines)
    s15 = re.search(r'临高县碧桂园金沙滩南享养生会所', lines)
    s16 = re.search(r'南享养生会所', lines)
    s17 = re.search(r'省定点医院', lines)
    s18 = re.search(r'临高县临城镇工会', lines)
    s19 = re.search(r'海口市美兰区三江农场派出所', lines)
    s20 = re.search(r'新媒体绿都小区', lines)
    s21 = re.search(r'美兰机场', lines)
    s22 = re.search(r'临高县半岛阳光小区', lines)
    s23 = re.search(r'三亚凤凰机场', lines)
    s24 = re.search(r'湖北襄阳机场', lines)
    s25 = re.search(r'明珠商业城百佳汇超市', lines)
    s26 = re.search(r'君澜三亚湾迎宾馆', lines)
    s27 = re.search(r'海口琼山大园', lines)
    s28 = re.search(r'文昌公园', lines)
    s29 = re.search(r'杏林小区旁瓜菜批发市场', lines)
    s30 = re.search(r'明珠商业城百佳汇超市', lines)
    s31 = re.search(r'海安港', lines)
    s32 = re.search(r'三亚市吉阳区君和君泰小区', lines)
    s33 = re.search(r'三亚市崖州区南滨佳康宾馆', lines)
    s34 = re.search(r'中西结合医院', lines)
    s35 = re.search(r'石碌镇城区', lines)
    s36 = re.search(r'昌化镇昌城村', lines)
    s37 = re.search(r'三亚定点留观点海角之旅酒店', lines)
    s38 = re.search(r'万宁市石梅湾九里三期', lines)
    s39 = re.search(r'三亚市吉阳区凤凰路南方航空城', lines)
    s40 = re.search(r'白马井镇海花岛', lines)

    lac_result = lac.run(lines)
    my_list = ['天涯海角景区', '海安', '旧港', '海安', '新港', '八爪鱼超市', '天涯区', '海口市', '板桥', '土福湾', '顺泽福湾', '亚龙湾壹号', '百花谷',
               '亚龙湾', '红塘湾建国酒店', '国光毫生度假酒店', '海棠区', '南田农场', '广西村', '水蛟村委会大园村', '新城路鲁能三亚湾升升超市', '金中海蓝钻小区'
        , '解放路', '旺豪超市', '南海', '嘉园', '广西自治区', '黄姚古镇', '一心堂', '大广岛药品超市', '钟廖村', '山根镇', '水央屈村', '万宁市'
        , '神州半岛', '君临海小区', '万城镇山泉海小区', '崖州区', '南海', '嘉园', '湛江市', '徐闻', '海安港', '南滨佳康宾馆', '崖州区', '东合逸海郡'
        , '云海台小区', '崖城怡宜佳超市', '吉阳区', '双大山湖湾小区', '星域小区', '神州租车公司', '水蛟村', '武汉市', '武汉市', '汉口站', '广西'
        , '南宁', '北海', '义龙西路汉庭酒店新温泉', '乐东', '千家镇', '只文村', '临城路', '亿佳超市', '君澜三亚湾', '迎宾馆', '儋州市', '光村镇'
        , '文昌市', '月亮城', '新海港', '琼海市', '琼海', '嘉积镇', '天来泉二区', '陵水县', '陵水', '英州镇', '椰田古寨', '兴隆镇', '曼特宁温泉'
        , '神州半岛', '君临海小区', '神州半岛', '新月海岸小区', '临高县', '碧桂园', '金沙滩', '浪琴湾', '儋州市', '光村镇', '解放军总医院', '海南医院'
        , '儋州', '客来湾海鲜店', '新港', '龙华区', '东方市', '涛升国际小区', '涛昇国际小区', '广东', '徐闻海安', '文昌', '东郊椰林', '秀英港'
        , '美兰机场', '石碌镇', '昌化镇', '澄迈县', '福山镇', '昌江县', '昌化镇', '昌城村', '三亚市', '东方市', '八所镇', '海口鹏', '泰兴购物广场'
        , '泰兴购物广场海甸三西路', '文化南路', '金洲大厦', '金廉路广电家园', '美兰区三江农场派出所', '蓝天路', '金廉路广电家园', '秀英区', '锦地翰城'
        , '定安县', '香江丽景小区', '石头公园', '美兰区', '海口市琼山区', '金宇街道', '城西镇', '中沙路', '琼山', '灵山镇', '琼山区', '新大洲大道'
        , '昌洒镇', '文城', '澄迈县', '老城镇', '万达广场', '湘忆亭', '骑楼小吃街', '龙桥镇', '玉良村', '老城开发区', '软件园商业步行街', '石山镇'
        , '美鳌村', '那大镇', '儋州街', '三亚市崖州区', '南滨佳康宾馆', '水蛟村委会', '三亚旺豪超市', '福源小区', '新三得羊羔店', '石碌镇晨鑫饭店'
        , '湖北', '湖北省', '武汉市', '湖南', '衡山', '新疆', '琼山区佳捷精品酒店', '滨江店', '江苏', '无锡', '福架村', '皇马假日游艇度假酒店'
        , '坡博农贸市场', '旺佳旺超市南沙店', '凤凰路南方航空城', '迎宾路', '南山寺']
    # my_list表示需要进行拼接的地名
    my1_list = ['南山寺', '天涯海角景区', '百花谷', '亚龙湾', '威斯汀酒店', '云海台小区', '崖城怡宜佳超市', '万应堂大药房购药', '天涯区', '文昌市月亮城',
                '千百汇超市', '乐东县人民医院', '三亚市人民医院', '君澜三亚湾迎宾馆', '云海台', '崖城怡宜佳超市', '海安新港', '杏林小区旁瓜菜批发市场',
                '文昌公园', '杏林小区旁瓜菜批发市场', '明珠商业城百佳汇超市', '一心堂', '大广岛药品超市', '南滨农贸市场', '南海嘉园', '三亚市吉阳区'
                                                                                         '新风街', '三亚湾', '凤翔路', '中山南路',
                '老城四季康城', '绿水湾小区', '天涯海角', '玫瑰谷', '福山骏群超市', '那大镇兰洋路', '中兴大道']
    # my1_list表示需要用分号隔开的地名
    location_list = []
    seg_list = lac_result[0]
    pos_list = lac_result[1]
    for i in range(len(seg_list)):
        if pos_list[i] == "LOC" or pos_list[i] == "ORG":
            location_list.append(seg_list[i])
    lines = lines.replace(' ', '')
    Tcode = re.search(
        r'(出租车|网约车|动车|私家车|旅游大巴|酒店专车|飞机|公交车|商务车|轮渡|旅游团车|滴滴专车|摩托车|救护车|急救车|海汽快车|高铁|电动车|航班|自驾|私家车|船|列车|骑车|驾车|火车|大巴|摩托三轮车|三轮车|小轿车|车|电火车|滴滴车|海汽专线|班车)',
        lines)  # 得到交通工具
    T = re.search(
        '([京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼使领][A-HJ-NP-Z][A-HJ-NP-Z0-9]{4}[A-HJ-NP-Z0-9挂学警港澳])|(班次号\s?[0-9]+)|(双泰\d+号|黎母号|海棠湾号|南方\d号轮渡|紫荆\d+号|粤海铁3号|黎姆号|海装18号)|(\d+)班车|(?:车次|火车)\s?([GCDZTSPKXLY1-9]\d{1,4})|乘([GCDZTSPKXLY1-9]\d{1,4})次火车|(?:国航|南航|海航|航班|航班号|航班号是)\s?([A-Z\d]{2}\d{3,4})|([A-Z\d]{2}\d{3,4})(?:次航|航班)|(?:乘|由|被)\s?([\d]+)(?=路|救护车|急救车|车)',
        lines)
    start = ''
    end = ''
    if s != None:
        end = s.group()
    if s1 != None:
        end = s1.group()
    if s2 != None:
        end = s2.group()
    if s3 != None:
        end = s3.group()
    if s4 != None:
        end = s4.group()
    if s5 != None:
        end = s5.group()
    if s6 != None:
        end = s6.group()
    if s7 != None:
        end = s7.group()
    if s8 != None:
        end = s8.group()
    if s9 != None:
        end = s9.group()
    if s10 != None:
        end = s10.group()
    if s11 != None:
        end = s11.group()
    if s12 != None:
        end = s12.group()
    if s13 != None:
        end = s13.group()
    if s14 != None:
        end = s14.group()
    if s15 != None:
        end = s15.group()
    if s16 != None:
        end = s16.group()
    if s17 != None:
        end = s17.group()
    if s18 != None:
        end = s18.group()
    if s19 != None:
        end = s19.group()
    if s20 != None:
        end = s20.group()
    if s21 != None:
        end = s21.group()
    if s22 != None:
        end = s22.group()
    if s23 != None:
        end = s23.group()
    if s24 != None:
        end = s24.group()
    if s25 != None:
        end = s25.group()
    if s26 != None:
        end = s26.group()
    if s27 != None:
        end = s27.group()
    if s28 != None:
        end = s28.group()
    if s29 != None:
        end = s29.group()
    if s30 != None:
        end = s30.group()
    if s31 != None:
        end = s31.group()
    if s32 != None:
        end = s32.group()
    if s33 != None:
        end = s33.group()
    if s34 != None:
        end = s34.group()
    if s35 != None:
        end = s35.group()
    if s36 != None:
        end = s36.group()
    if s37 != None:
        end = s37.group()
    if s38 != None:
        end = s38.group()
    if s39 != None:
        end = s39.group()
    if s40 != None:
        end = s40.group()
    if len(location_list) == 2:
        if location_list[0] in my_list and location_list[1]:
            end = location_list[0] + location_list[1]
            start = ''  # 如果识别到的两个地名，都是在my_list里面，就相当于原先只有一个地名，那么被切分成两个地名
            # ，所以我们需要把这两个地名合起来作为终点，终点就相当于两者的拼接
        else:
            start = location_list[0]
            end = location_list[1]  # 否则的话，就是这两个地点不是由一个地点拆分的，那么我们就
            # 将第一个作为起点，然后将第二个起点作为终点
        if location_list[0] in my1_list and location_list[1] in my1_list:
            end = location_list[0] + ';' + location_list[1]
            start = ''  # 如果识别到的两个地点都是在my_list里面的话，那么意思就是两个地点都为终点，所以我们需要将这两个地点存入终点中
            # 并且用分号隔开
    if len(location_list) == 3:
        if location_list[0] in my_list and location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[0] + location_list[1] + location_list[2]
            start = ''
        elif location_list[0] in my1_list and location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[1] + location_list[2]
            start = location_list[0] + ';' + end
        elif location_list[0] in my_list and location_list[1] in my_list:
            start = location_list[0] + location_list[1]
            end = location_list[2]
        elif location_list[1] in my_list and location_list[2] in my_list:
            end = location_list[1] + location_list[2]
            start = location_list[0]
    if T != None and Tcode:  # 判断是否坐的飞机，因为只有飞机有航班号
        # 提取车牌号
        if T.group(1):
            t = T.group(1)

        # 提取班次号
        if T.group(2):
            t = T.group(2)

        # 提取特殊交通  (双泰\d+号|黎母号|海棠湾号|南方\d号轮渡|紫荆\d+号|粤海铁3号|黎姆号|海装18号)
        if T.group(3):
            t = T.group(3)

        # 提取班车号
        if T.group(4):
            t = T.group(4)

        # 提取火车有两种格式
        if T.group(5):
            t = T.group(5)
        elif T.group(6):
            t = T.group(6)

        # 提取航班号有两种格式
        if T.group(7):
            t = T.group(7)
        elif T.group(8):
            t = T.group(8)

        # 提取特殊交通 例如120
        if T.group(9):
            t = T.group(9)

        dict_y = [t, Tcode.group(), start, end]
        list1.extend(dict_y)

    elif Tcode:

        dict_y = ['', Tcode.group(), start, end]
        list1.extend(dict_y)
    elif T:
        dict_y = [T.group(), '', start, end]
        list1.extend(dict_y)
    else:
        dict_y = ['', '', start, end]
        list1.extend(dict_y)
    return list1


def allin():
    # 导入excel文档
    wb = openpyxl.load_workbook('./merged.xlsx')
    # 链接到航班数据库，host为本地主机，端口是3306，用户是root，密码是Scott624，连接上的数据库为flight，编码是utf8
    link = sqlite3.connect('transport.db')

    # 获取游标对象，游标是系统为用户开设的一个数据缓冲区，存放SQL语句的执行结果，结果是零条、一条或多条数据
    data1 = link.cursor()
    data2 = link.cursor()
    data3 = link.cursor()
    data4 = link.cursor()
    write1 = link.cursor()
    write2 = link.cursor()

    # 获取excel文档的数据表
    ws = wb['数据']

    # 命名要存放数据的列
    ws['N1'] = '机场/火车站点'
    ws['O1'] = '火车站序'
    ws['P1'] = '出发时间'
    ws['Q1'] = '到达时间'
    # 获取excel数据表的，第I列单元格
    colI = ws['J']
    # 获取I列单元格的最大长度
    _max = len(colI)
    # 因为第一行数据是不需要进行处理的，所以处理数据要从第二行开始
    x = 2

    # 从第九列的第二行到最后一行，读取数据，读取的数据是列车号或者航班号
    for i in ws.iter_rows(min_row=2, max_row=_max, max_col=10, min_col=10):
        for _cell in i:
            # 列车号或者航班号不为空
            if _cell.value != None:
                # 并且该行的第10列的内容为航班时
                if ws.cell(x, 11).value == '航班':

                    # 获取该行第K、L列的数据。第K列的数据是出发地，第L列的数据为目的地
                    StartStation_cell = 'L' + str(x)
                    ArriveStation_cell = 'M' + str(x)
                    # 获得单元格内容
                    StartStation = ws[StartStation_cell]
                    ArriveStation = ws[ArriveStation_cell]

                    # 使用sql语句查询数据库。EXECUTE语句可以执行存放SQL语句的字符串变量，或直接执行SQL语句字符串。需要查询的数据由.format来填充，StartStation.value是出发地，ArriveStation.value[0:2]取目的地的前两个字为查询输入字段，_cell.value是航班号

                    if StartStation.value != None and ArriveStation.value != None:
                        data1.execute(
                            "SELECT StartDrome,ArriveDrome,StartTime,ArriveTime from tbl_air where startCity=(select a.Abbreviation from tbl_icao as a where a.Address_cn='{}') and lastCity=(select b.Abbreviation from tbl_icao as b where b.Address_cn='{}') and AirlineCode='{}';".format(
                                StartStation.value, ArriveStation.value[0:2], _cell.value))

                        # fetchall使用游标从查询中检索
                        All = data1.fetchall()
                        # 如果All中的数据不为空
                        if All:
                            # 首先利用sql语句，将数据写入一个新表中。_cell.value是航班号，All[0][0]是出发机场，All[0][1]是降落机场，All[0][2]是起飞时间，All[0][3]是降落时间。同时要在起飞机场、降落机场和航班号都不同的情况下才写入到新的表中
                            write1.execute(
                                "insert or ignore into tbl_airnum (Code,Start, Arrive, StartTime, ArriveTime) values ('{}','{}','{}', '{}','{}');".format(
                                    _cell.value, All[0][0], All[0][1], All[0][2], All[0][3]))
                            # 表示允许写入
                            link.commit()

                            # All中的数据，分别存放的是起飞机场、降落机场、起飞时间和降落时间

                            # 用于存放起降机场
                            row1 = 'N' + str(x)
                            # 用于存放起飞时间
                            row2 = 'P' + str(x)
                            # 用于存放降落时间
                            row3 = 'Q' + str(x)

                            # 起飞机场和降落机场之间由'|'隔开
                            ws[row1] = All[0][0] + '|' + All[0][1]
                            # 写入起飞时间
                            ws[row2] = All[0][2]
                            # 写入降落时间
                            ws[row3] = All[0][3]

                    # 保存为8.xlsx
                    wb.save('22.xlsx')

                else:
                    # 用于存放列车经过的站点
                    Save_station = ""
                    # 用于存放列车经过站点的站序
                    Save_S_No = ""
                    # 用于存放列车的进站时间
                    Save_A_Time = ""
                    # 用于存放列车的出站时间
                    Save_D_Time = ""

                    if _cell.value != None:
                        # 获取该行第K、L列的数据。第K列的数据是出发地，第L列的数据为目的地
                        Train_Start_cell = 'L' + str(x)
                        Train_Arrive_cell = 'M' + str(x)

                        Train_Start = ws[Train_Start_cell]
                        Train_Arrive = ws[Train_Arrive_cell]

                        # 利用sql语句进行查询。通过列车号、出发站和到达站查询途径站点、站序、进站时间和出站时间。_cell.value是列车号，Train_Start.value是出发站，Train_Arrive.value是到达站
                        data2.execute(
                            "select Station,S_No,A_Time,D_Time from tbl_train where ID = '{}' and S_No >= (select S_No from tbl_train where ID = '{}' and Station like '{}%' limit 1) and S_No <= (select S_No from tbl_train where ID = '{}' and Station like '{}%' limit 1) order by S_No;".format(
                                _cell.value, _cell.value, Train_Start.value, _cell.value, Train_Arrive.value))

                        # result中的数据分别是站名、站序、到达时间和出发时间
                        result = data2.fetchall()

                        # 获取result的最后一个元素的索引下标
                        Tlen = len(result) - 1

                        for i in range(len(result)):
                            # 将所经过的站点用'|'隔开，并在每一个站后加上一个'站'字
                            Save_station = Save_station + result[i][0] + '站' + '|'
                            # 将所经过的站点的站序用'|'隔开
                            Save_S_No = Save_S_No + str(result[i][1]) + '|'
                            # 将所经过的站点的进站时间用'|'隔开
                            Save_A_Time = Save_A_Time + result[i][2] + '|'
                            # 将所经过的站点的出站时间用'|'隔开
                            Save_D_Time = Save_D_Time + result[i][3] + '|'

                        # 将所有需要储存的数据的最后一个竖线去除
                        Save_station = Save_station[0:len(Save_station) - 1]
                        Save_S_No = Save_S_No[0:len(Save_S_No) - 1]
                        Save_A_Time = Save_A_Time[0:len(Save_A_Time) - 1]
                        Save_D_Time = Save_D_Time[0:len(Save_D_Time) - 1]

                        # 如果result中的存有数据
                        if result:
                            # 将列车号、出发站的站序和到达站的站序储存在新表中。_cell.value是列车号，result[0][1]是出发站的站序，result[Tlen][1]是到达站的站序
                            write2.execute(
                                "insert or ignore into tbl_trainum (Code,startNum,arriveNum) values ('{}','{}','{}');".format(
                                    _cell.value, result[0][1], result[Tlen][1]))
                            link.commit()

                            # 储存站名
                            row1 = 'N' + str(x)
                            # 储存站序
                            row2 = 'O' + str(x)
                            # 储存到达时间
                            row3 = 'P' + str(x)
                            # 储存发车时间
                            row4 = 'Q' + str(x)

                            ws[row1] = Save_station
                            ws[row2] = Save_S_No
                            ws[row3] = Save_A_Time
                            ws[row4] = Save_D_Time

                        wb.save('22.xlsx')

            x = x + 1

    # 最后关闭连接
    data1.close()
    data2.close()
    data3.close()
    data4.close()
    link.close()


"""
功能： 读取Excel地点数据请求经纬度信息存储Excel文件
"""


def write2Excel(result, save_path="result.xls"):
    """
    存入Excel
    """
    j = 0
    book = Workbook(encoding="utf-8")
    table = book.add_sheet("result")
    for one_list in result:
        for i in range(len(one_list)):
            table.write(j, i, one_list[i])
        j += 1
    book.save(save_path)


def readExcel(data="22.xlsx"):
    """
    加载Excel数据表
    """
    workbook = xlrd.open_workbook(data, encoding_override="utf-8")
    table = workbook.sheets()[0]
    row_num, col_num = table.nrows, table.ncols
    data_list = []
    for i in range(row_num):
        one_list = []
        for j in range(col_num):
            one_list.append(table.cell_value(i, j))
        data_list.append(one_list)
    return data_list


def singleRequest(location="中山大学", limit_city=""):
    """
    单次请求处理
    """
    params = {
        "ak": "H1r9OqNNZb2w92ezkhpDboOoXHkGuxKH",
        "region": "null" if limit_city == "" else limit_city,
        "output": "json",
        "query": location,
    }
    query_url = f"https://api.map.baidu.com/place/v2/suggestion?{urlencode(params)}"
    res = requests.get(query_url).json()
    lat = "None"
    lng = "None"
    name = location
    try:
        lat = res["result"][0]["location"]["lat"]
        lng = res["result"][0]["location"]["lng"]
        name = res["result"][0]["name"]
    except:
        pass
    res = ",".join([str(O) for O in [lat, lng, name]])
    return res


def main(data_path="22.xlsx", save_path="result.xlsx"):
    """
    主函数
    """
    data_list = readExcel(data=data_path)
    res_list = []
    title = data_list[0]
    title += ["出发点经纬度", "目的地经纬度"]
    res_list.append(title)
    for one_list in data_list[1:]:
        start, end = one_list[11], one_list[12]
        Q = singleRequest(location=start)
        R = singleRequest(location=end)
        one_list += [Q, R]
        res_list.append(one_list)
        time.sleep(random.uniform(0.1, 0.8))
    write2Excel(res_list, save_path=save_path)


if __name__ == '__main__':
    # filepath = input('请输入文件路径：')
    # LInes = load_xlsx("CRF_Hainan.xlsx")
    # Save_to_excle("hainan_Track15.xlsx", LInes)
    # SplitText("hainan_Track15.xlsx", "SPlitByStop.txt", "SplitTextUltimate.txt", "hainan_Track15.xlsx")#数据预处理
    data_handle('./hainan_Track15.xlsx')  # 提取文档的内容
    merge(extend_file="./excel.xlsx", res_file="海南疫情.xlsx")
    # allin()
    # main(data_path="./22.xlsx", save_path="result.xlsx")
